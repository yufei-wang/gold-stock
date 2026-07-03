# -*- coding: utf-8 -*-
"""
金股动态网站数据构建脚本
用法: python build_site.py
功能: 扫描上级目录中所有总结Excel，解析后输出 data.js 供 HTML 使用
"""

import os
import re
import json
from openpyxl import load_workbook


# ========== 配置 ==========
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.js')


def extract_month_from_filename(filename):
    """从文件名中提取年月, 如 20260703 -> (2026, 7)"""
    m = re.search(r'(\d{4})(\d{2})\d{2}', filename)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def parse_summary_sheet(ws):
    """解析总结文件中的汇总集合表"""
    records = []
    max_row = ws.max_row
    max_col = ws.max_column
    if not max_row or max_row < 2:
        return records

    # 读取表头
    headers = {}
    for c in range(1, min(max_col + 1, 20)):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c

    industry_col = headers.get('一级行业')
    code_col = headers.get('证券代码')
    name_col = headers.get('证券名称')
    count_col = headers.get('推荐次数')
    analyst_col = headers.get('券商/分析师')
    reason_col = headers.get('推荐理由')

    # 找变化列（列名含"变化"）
    change_col = None
    for h, c in headers.items():
        if '变化' in h:
            change_col = c
            break

    for r in range(2, max_row + 1):
        name_val = ws.cell(row=r, column=name_col).value if name_col else None
        if not name_val:
            continue
        name = str(name_val).strip()
        if name in ('None', 'nan', ''):
            continue

        code_val = ws.cell(row=r, column=code_col).value if code_col else ''
        code = str(code_val).strip() if code_val else ''

        industry_val = ws.cell(row=r, column=industry_col).value if industry_col else ''
        industry = str(industry_val).strip() if industry_val else ''
        if industry in ('None', 'nan'):
            industry = ''

        count_val = ws.cell(row=r, column=count_col).value if count_col else 0
        try:
            count = int(count_val)
        except (ValueError, TypeError):
            count = 0

        analyst_val = ws.cell(row=r, column=analyst_col).value if analyst_col else ''
        analyst = str(analyst_val).strip() if analyst_val else ''
        if analyst in ('None', 'nan'):
            analyst = ''

        change_val = ws.cell(row=r, column=change_col).value if change_col else ''
        change = str(change_val).strip() if change_val is not None else ''
        if change in ('None', 'nan'):
            change = ''

        reason_val = ws.cell(row=r, column=reason_col).value if reason_col else ''
        reason = str(reason_val).strip() if reason_val is not None else ''
        if reason in ('None', 'nan'):
            reason = ''

        records.append({
            'industry': industry,
            'code': code,
            'name': name,
            'count': count,
            'analysts': analyst,
            'change': change,
            'reason': reason,
        })

    return records


def load_prices():
    """加载 prices.json（若存在）"""
    price_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'prices.json')
    if not os.path.exists(price_file):
        return None
    try:
        with open(price_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"警告: 读取 prices.json 失败: {e}")
        return None


# 各月首个交易日（用于查基准价）
MONTH_BASE_DATE = {
    '2026-05': '20260506',
    '2026-06': '20260601',
    '2026-07': '20260701',
}


def scan_and_build():
    """扫描目录，构建全部数据"""
    data = {
        'months': [],
        'summary': {},
        'prices_meta': None,
    }

    # 扫描总结文件
    files = os.listdir(DATA_DIR)
    summary_files = sorted([f for f in files if '总结' in f and f.endswith('.xlsx')])

    print(f"找到 {len(summary_files)} 个总结文件")

    for fname in summary_files:
        year, month = extract_month_from_filename(fname)
        if not year:
            continue
        month_key = f"{year}-{month:02d}"
        print(f"  处理: {fname} -> {month_key}")

        fpath = os.path.join(DATA_DIR, fname)
        wb = load_workbook(fpath, read_only=True, data_only=True)

        for sn in wb.sheetnames:
            if '汇总' in sn and '集合表' in sn:
                records = parse_summary_sheet(wb[sn])
                if records:
                    data['summary'][month_key] = records
                    print(f"    -> {sn}: {len(records)} 条")
                break
        wb.close()

    # 月份列表
    data['months'] = sorted(data['summary'].keys())

    # ---- 合并价格数据（仅最新3个月） ----
    prices_data = load_prices()
    if prices_data:
        prices = prices_data.get('prices', {})
        latest_date = prices_data.get('latest_date', '20260703')
        data['prices_meta'] = {
            'generated_at': prices_data.get('generated_at'),
            'latest_date': latest_date,
        }
        # 只为最近1个月合并涨幅
        target_months = data['months'][-1:]
        print(f"\n为最新月份计算涨幅: {target_months}")
        matched = 0
        for mk in data['months']:
            base_date = MONTH_BASE_DATE.get(mk)
            for r in data['summary'][mk]:
                code = r.get('code', '')
                # 早期月份不计算涨幅
                if mk not in target_months or not base_date:
                    r['base_price'] = None
                    r['latest_price'] = None
                    r['change_pct'] = None
                    r['currency'] = None
                    r['base_date'] = None
                    continue

                p = prices.get(code)
                if not p:
                    r['base_price'] = None
                    r['latest_price'] = None
                    r['change_pct'] = None
                    r['currency'] = None
                    r['base_date'] = base_date
                    continue

                base_price = p.get(base_date)
                latest_price = p.get(latest_date)
                r['base_price'] = base_price
                r['latest_price'] = latest_price
                r['currency'] = p.get('currency', 'CNY')
                r['base_date'] = base_date
                if base_price and latest_price:
                    r['change_pct'] = round((latest_price - base_price) / base_price * 100, 2)
                    matched += 1
                else:
                    r['change_pct'] = None
        print(f"价格数据已合并: {matched} 条记录含涨幅")
    else:
        print("\n未找到 prices.json，跳过价格合并")

    return data


def main():
    print("=" * 50)
    print("金股动态网站 - 数据构建")
    print(f"数据目录: {DATA_DIR}")
    print(f"输出文件: {OUTPUT_FILE}")
    print("=" * 50)

    data = scan_and_build()

    js_content = f"// 自动生成，请勿手动编辑\n// 生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\nconst GOLD_STOCK_DATA = {json.dumps(data, ensure_ascii=False, indent=2)};\n"

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f"\n数据已输出到: {OUTPUT_FILE}")
    print(f"共 {len(data['months'])} 个月份: {', '.join(data['months'])}")
    for mk in data['months']:
        print(f"  {mk}: {len(data['summary'].get(mk, []))} 条")


if __name__ == '__main__':
    main()
